from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from zipfile import BadZipFile
import os

FILE_PATH = "data/summaries.xlsx"

def save_to_excel(rows):
    if os.path.exists(FILE_PATH):
        try:
            wb = load_workbook(FILE_PATH)
            ws = wb.active
        except BadZipFile:
            os.remove(FILE_PATH)
            wb = Workbook()
            ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 110

    for title, url, summary in rows:
        ws.append(["TITLE", title])
        ws.append(["URL", url])
        ws.append(["SUMMARY", summary])

        author = ""
        publisher = ""

        for line in summary.splitlines():
            if line.lower().startswith("author:"):
                author = line.replace("Author:", "").strip()
            if line.lower().startswith("publisher:"):
                publisher = line.replace("Publisher:", "").strip()

        if author:
            ws.append(["Author", author])
        if publisher:
            ws.append(["Publisher", publisher])

        ws.append(["", ""])

    for row in ws.iter_rows(min_col=1, max_col=2):
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(FILE_PATH)
